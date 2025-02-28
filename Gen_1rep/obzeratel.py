import ast
import logging
import os
import pandas as pd
import numpy as np
from matplotlib import pyplot as plt
from openpyxl import *
from openpyxl.chart import ScatterChart, Reference
from openpyxl.chart.series import Series
from openpyxl.drawing.image import Image
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter


def find_matching_excel_files(root_dir):
    matching_files = []

    for dirpath, dirnames, filenames in os.walk(root_dir):
        for dirname in dirnames:
            # Формируем полное имя файла, которое должно совпадать с названием папки
            expected_file = os.path.join(dirpath, dirname, f"{dirname}.xlsx")

            # Проверяем, существует ли такой файл
            if os.path.exists(expected_file):
                matching_files.append(expected_file)

    return matching_files

def safe_eval(cell):
    if isinstance(cell, str):
        # Замена запятой на точку и nan на None перед преобразованием
        cell = cell.replace('nan', 'None')
        try:
            return ast.literal_eval(cell)
        except (SyntaxError, ValueError):
            return cell
    return cell

def tuple_to_float(value):
    if isinstance(value, float):
        return value  # Если уже float, возвращаем как есть
    elif isinstance(value, tuple) and len(value) == 2:
        try:
            # Преобразуем кортеж в строку вида '1.5'
            tuple_str = str(value[0]) + '.' + str(value[1])
            # Преобразуем строку в число с плавающей точкой
            return float(tuple_str)
        except (ValueError, TypeError):
            return None  # Возвращаем None при ошибке
    else:
        return value

def convert_dict_to_list(cell):
    if isinstance(cell, dict):
        return list(cell.values())
    return cell

def flatten_list(cell):
    if isinstance(cell, list):
        return ', '.join(map(str, cell))
    return cell


def remove_duplicates_and_none(s):
    if pd.isna(s):
        return ''
    values = s.split(',')
    cleaned_values = []
    for value in values:
        value = value.strip()
        if value != 'None' and value not in cleaned_values:
            cleaned_values.append(value)
    return ', '.join(cleaned_values)

def Picator(d_f, ax_x, ax_y, x_nam, y_nam, indicator, graf_nam, sheet_nam):
    # Функция для безопасного преобразования
    def convert_to_list(x):
        try:
            if isinstance(x, str):
                # Разбиваем строку и преобразуем каждый элемент
                return [float(i) if i.lower() != 'none' else None for i in x.split(', ')]
            elif isinstance(x, (float, int)):
                return [x]
            else:
                return []
        except ValueError as e:
            logging.warning(f"Ошибка при преобразовании: {x}, ошибка: {e}")
            return []

    # Применяем функцию к столбцам
    d_f[ax_x] = d_f[ax_x].apply(
        convert_to_list)

    d_f[ax_y] = d_f[ax_y].apply(
        lambda x: convert_to_list(x) if isinstance(x, str) else
            [float(x)] if pd.notna(x) else [])

    # Остальной код остается прежним:
    d_f = d_f.explode([ax_x, ax_y])
    d_f = d_f.dropna(subset=[ax_y])
    d_f = d_f.reset_index(drop=True)

    # Создаем график
    plt.figure(figsize=(12, 8))

    # Группируем по лабораториям
    for lab, group in d_f.groupby(indicator):
        plt.plot(group[ax_x],
                 group[ax_y],
                 'o',
                 label=lab)

    plt.title(graf_nam)
    plt.xlabel(x_nam)
    plt.ylabel(y_nam)
    plt.legend(loc='upper right')
    plt.grid(True)
    # plt.show()


    # Сохраняем график как изображение
    plt.savefig(f'{sheet_nam}.jpg', bbox_inches='tight', dpi=300)

    # Открываем существующий файл Excel
    excel_file = output_path
    sheet_name = sheet_nam

    # Загружаем существующий файл
    book = load_workbook(excel_file)
    sheet = book[sheet_name]

    # Загружаем изображение
    img = Image(f'{sheet_nam}.jpg')

    # Настраиваем размер и положение изображения
    img.width = 600  # в пикселях
    img.height = 400  # в пикселях
    img.anchor = 'D2'  # начальная ячейка для вставки

    # Добавляем изображение на лист
    sheet.add_image(img)

    # Сохраняем изменения
    book.save(excel_file)




def DfList(initial_rout, text_to_remove, column_to_remove, list_mes_transformation, duplicat_list, time_column):
    files = find_matching_excel_files(initial_rout)
    dataframes = []
    for file in files:
        df = pd.read_excel(file, index_col=0)

        df_transposed = df.transpose()
        values_to_remove = text_to_remove
        df_transposed = df_transposed.replace(values_to_remove, np.nan)
        df_transposed = df_transposed.map(safe_eval)
        df_transposed = df_transposed.map(convert_dict_to_list)
        df_transposed = df_transposed.map(flatten_list)
        df_transposed.columns = df_transposed.iloc[0]
        df_transposed = df_transposed.drop(df_transposed.index[0])
        df_transposed = df_transposed.drop(column_to_remove, axis=1)

        tuple_list = list_mes_transformation
        for t_l in tuple_list:
            df_transposed[t_l] = df_transposed[t_l].apply(tuple_to_float)
        clean_list = duplicat_list
        for c_l in clean_list:
            df_transposed[c_l] = df_transposed[c_l].apply(remove_duplicates_and_none)
        try:
            time_col = time_column
            for col in time_col:
                df_transposed[col] = pd.to_datetime(df_transposed[col]).dt.strftime('%d.%m.%Y')
        except:
            pass

        dataframes.append(df_transposed)
    return dataframes

def MergDF(df_list, index_ind, file_nam, sheet_nam):
    merged_df = pd.concat(df_list, ignore_index=True)
    # merged_df = merged_df.map(type)
    m_df = merged_df.set_index(index_ind)
    s_df = m_df.sort_index()
    # output_path = r'C:\Users\epyur\PycharmProjects\PythonProject\Gen_1rep\rep'  # Замените на путь к выходной папке
    current_directory = os.getcwd()
    output_path = os.path.join(current_directory, 'rep', f'{file_nam}.xlsx')
    writer = pd.ExcelWriter(output_path, engine='xlsxwriter')
    s_df.to_excel(writer, index=True, sheet_name=sheet_nam)
    return s_df


workbook = writer.book
worksheet = writer.sheets['Результаты_испытаний']

format_work = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})  # Красный
worksheet.conditional_format('A1:CT' + str(len(s_df)+1), {
 'type': 'text',
 'criteria': 'containing',
 'value': 'Не соответствует',
 'format': format_work
})

format_w = workbook.add_format({'bg_color': '#00FF00', 'font_color': '#000000'})  # Красный
worksheet.conditional_format('A1:CT' + str(len(s_df)+1), {
 'type': 'text',
 'criteria': 'containing',
 'value': 'Соответствует',
 'format': format_w
})


# Сохраняем файл
writer.close()

def ConditionFrame(d_f, condition, sheet_nam):
    with pd.ExcelWriter(output_path, mode='a', engine='openpyxl') as writer:
        new_df = d_f[condition]
        new_df.to_excel(writer, sheet_name=sheet_nam)
    return new_df



def SelectAndFiltre(d_f, select_column, column_name, filter, sheet_nam):
    with pd.ExcelWriter(output_path, mode='a', engine='openpyxl') as writer:
        filtered_data = merged_df[select_column]
        filtered_data = filtered_data[filtered_data[column_name] == filter]
        new_dataframe = filtered_data.copy()
        new_dataframe.to_excel(writer, sheet_name=sheet_nam)
    return new_dataframe


