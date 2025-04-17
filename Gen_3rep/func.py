import ast
import logging
import os
from datetime import datetime
from pathlib import Path
import pandas as pd
import numpy as np
from matplotlib import pyplot as plt
from openpyxl import *
from openpyxl.chart import ScatterChart, Reference
from openpyxl.chart.series import Series
from openpyxl.drawing.image import Image
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter


def find_matching_excel_files(root_dir):
    matching_files = []

    for dirpath, dirnames, filenames in os.walk(root_dir):
        for dirname in dirnames:
            # Формируем полное имя файла, которое должно совпадать с названием папки
            expected_file = os.path.join(dirpath, dirname, f"{dirname}.xlsx")

            # Проверяем, существует ли такой файл
            if os.path.exists(expected_file):
                matching_files.append(expected_file)

    return matching_files

def safe_eval(cell):
    if isinstance(cell, str):
        # Замена запятой на точку и nan на None перед преобразованием
        cell = cell.replace('nan', 'None')
        try:
            return ast.literal_eval(cell)
        except (SyntaxError, ValueError):
            return cell
    return cell

def tuple_to_float(value):
    if isinstance(value, float):
        return value  # Если уже float, возвращаем как есть
    elif isinstance(value, tuple) and len(value) == 2:
        try:
            # Преобразуем кортеж в строку вида '1.5'
            tuple_str = str(value[0]) + '.' + str(value[1])
            # Преобразуем строку в число с плавающей точкой
            return float(tuple_str)
        except (ValueError, TypeError):
            return None  # Возвращаем None при ошибке
    else:
        return value

def convert_dict_to_list(cell):
    if isinstance(cell, dict):
        return list(cell.values())
    return cell

def flatten_list(cell):
    if isinstance(cell, list):
        return ', '.join(map(str, cell))
    return cell


def remove_duplicates_and_none(s):
    if pd.isna(s):
        return ''
    values = s.split(',')
    cleaned_values = []
    for value in values:
        value = value.strip()
        if value != 'None' and value not in cleaned_values:
            cleaned_values.append(value)
    return ', '.join(cleaned_values)

def Picator(d_f, ax_x, ax_y, x_nam, y_nam, indicator, graf_nam, file_nam, sheet_nam):
    # Функция для безопасного преобразования
    def convert_to_list(x):
        try:
            if isinstance(x, str):
                # Разбиваем строку и преобразуем каждый элемент
                return [float(i) if i.lower() != 'none' else None for i in x.split(', ')]
            elif isinstance(x, (float, int)):
                return [x]
            else:
                return []
        except ValueError as e:
            logging.warning(f"Ошибка при преобразовании: {x}, ошибка: {e}")
            return []

    # Применяем функцию к столбцам
    d_f[ax_x] = d_f[ax_x].apply(
        convert_to_list)

    d_f[ax_y] = d_f[ax_y].apply(
        lambda x: convert_to_list(x) if isinstance(x, str) else
            [float(x)] if pd.notna(x) else [])

    d_f = d_f.explode([ax_x, ax_y])
    d_f = d_f.dropna(subset=[ax_y])
    d_f = d_f.reset_index(drop=True)

    # Создаем график
    plt.figure(figsize=(12, 8))

    # Группируем по лабораториям
    for lab, group in d_f.groupby(indicator):
        plt.plot(group[ax_x],
                 group[ax_y],
                 'o',
                 label=lab)

    plt.title(graf_nam)
    plt.xlabel(x_nam)
    plt.ylabel(y_nam)
    plt.legend(loc='upper right')
    plt.grid(True)
    # plt.show()


    # Сохраняем график как изображение
    plt.savefig(f'{sheet_nam}.jpg', bbox_inches='tight', dpi=300)

    # Открываем существующий файл Excel
    excel_file = file_nam
    sheet_name = sheet_nam

    # Загружаем существующий файл
    book = load_workbook(excel_file)
    sheet = book[sheet_name]

    # Загружаем изображение
    img = Image(f'{sheet_nam}.jpg')

    # Настраиваем размер и положение изображения
    img.width = 600  # в пикселях
    img.height = 400  # в пикселях
    img.anchor = 'P2'  # начальная ячейка для вставки

    # Добавляем изображение на лист
    sheet.add_image(img)

    # Сохраняем изменения
    book.save(excel_file)


def DfList(initial_rout, text_to_remove, column_to_remove, list_mes_transformation, duplicat_list, time_column):
    files = find_matching_excel_files(initial_rout)
    dataframes = []
    for file in files:
        df = pd.read_excel(file, index_col=0)

        df_transposed = df.transpose()
        values_to_remove = text_to_remove
        df_transposed = df_transposed.replace(values_to_remove, np.nan)
        df_transposed = df_transposed.map(safe_eval)
        df_transposed = df_transposed.map(convert_dict_to_list)
        df_transposed = df_transposed.map(flatten_list)
        df_transposed.columns = df_transposed.iloc[0]
        df_transposed = df_transposed.drop(df_transposed.index[0])
        df_transposed = df_transposed.drop(column_to_remove, axis=1)

        tuple_list = list_mes_transformation
        for t_l in tuple_list:
            df_transposed[t_l] = df_transposed[t_l].apply(tuple_to_float)
        clean_list = duplicat_list
        for c_l in clean_list:
            df_transposed[c_l] = df_transposed[c_l].apply(remove_duplicates_and_none)
        try:
            time_col = time_column
            for col in time_col:
                df_transposed[col] = pd.to_datetime(df_transposed[col]).dt.strftime('%d.%m.%Y')
        except:
            pass

        dataframes.append(df_transposed)
    return dataframes

def MergDF(df_list, index_ind, file_nam, sheet_nam):
    try:
        # Проверка входных данных
        if not isinstance(df_list, list):
            raise TypeError("df_list должен быть списком датафреймов")

        # Объединение датафреймов
        merged_df = pd.concat(df_list, ignore_index=True)

        # Проверка наличия столбца для индекса
        if index_ind not in merged_df.columns:
            raise ValueError(f"Столбец {index_ind} не найден в датафрейме")

        # Установка и сортировка индекса
        m_df = merged_df.set_index(index_ind)
        s_df = m_df.sort_index()

        # Создание директории, если не существует
        Path(file_nam).parent.mkdir(parents=True, exist_ok=True)

        # Экспорт в Excel
        with pd.ExcelWriter(file_nam, engine='xlsxwriter') as writer:
            s_df.to_excel(writer, index=True, sheet_name=sheet_nam)

        print(f"Файл успешно сохранен: {file_nam}")

    except FileNotFoundError as e:
        print(f"Ошибка: файл не найден - {e}")
    except ValueError as e:
        print(f"Ошибка: {e}")
    except TypeError as e:
        print(f"Ошибка: неверный тип данных - {e}")
    except Exception as e:
        print(f"Произошла ошибка: {e}")
    return s_df


def ColorMark(file_nam, sheet_nam, red_key, green_key):
    try:
        # Проверка существования файла
        if not Path(file_nam).exists():
            raise FileNotFoundError(f"Файл {file_nam} не найден")

        # Читаем существующий файл
        existing_df = pd.read_excel(file_nam, sheet_name=sheet_nam)

        # Создаем writer для перезаписи файла
        writer = pd.ExcelWriter(file_nam, engine='xlsxwriter')
        workbook = writer.book

        # Добавляем существующий датафрейм
        existing_df.to_excel(writer, sheet_name=sheet_nam, index=False, header=True)

        # Получаем существующий лист
        worksheet = writer.sheets[sheet_nam]

        # Форматирование для красного цвета
        format_red = workbook.add_format({
            'bg_color': '#FFC7CE',
            'font_color': '#9C0006'
        })

        # Форматирование для зеленого цвета
        format_green = workbook.add_format({
            'bg_color': '#00FF00',
            'font_color': '#000000'
        })

        # Установка области форматирования
        last_row = len(existing_df) + 1
        format_range = f'A1:CT{last_row}'

        # Условие для красного цвета
        worksheet.conditional_format(
            format_range,
            {
                'type': 'text',
                'criteria': 'containing',
                'value': red_key,
                'format': format_red
            }
        )

        # Условие для зеленого цвета
        worksheet.conditional_format(
            format_range,
            {
                'type': 'text',
                'criteria': 'containing',
                'value': green_key,
                'format': format_green
            }
        )

        # Сохраняем изменения
        writer.close()
        print(f"Форматирование успешно применено к файлу {file_nam}")

    except FileNotFoundError as e:
        print(f"Ошибка: {e}")
    except KeyError as e:
        print(f"Ошибка: лист {sheet_nam} не найден")
    except Exception as e:
        print(f"Произошла ошибка: {e}")

    except FileNotFoundError as e:
        print(f"Ошибка: {e}")
    except KeyError as e:
        print(f"Ошибка: лист {sheet_nam} не найден")
    except Exception as e:
        print(f"Произошла ошибка: {e}")

def ConditionFrame(d_f, condition, file_nam, sheet_nam):
    with pd.ExcelWriter(file_nam, mode='a', engine='openpyxl') as writer:
        new_df = d_f[condition]
        new_df.to_excel(writer, sheet_name=sheet_nam)
    return new_df


def SelectAndFilter(d_f,
                    select_column = None,
                    column_name=None,
                    filters=None,
                    date_column=None,
                    file_nam=None,
                    sheet_nam=None,
                    dict_to_rename=None,
                    start_date=None,
                    end_date=None,
                    quarter=None,
                    half_year=None,
                    drop_date=None):
    try:
        # Проверка на None перед работой с датафреймом
        if d_f is None:
            raise ValueError("Входной датафрейм не должен быть None")

        d_f = d_f.reset_index()
        if select_column:
            # Основная фильтрация
            filtered_data = d_f[select_column]
        else:
            filtered_data = d_f

        if column_name:
            filtered_data = filtered_data[filtered_data[column_name].isin(filters)]

        # Конвертация столбца в формат даты
        if date_column:
            filtered_data[date_column] = pd.to_datetime(filtered_data[date_column], format='%d.%m.%Y', errors='coerce')

        # Фильтрация по дате
        if start_date:
            start_date = datetime.strptime(start_date, '%d.%m.%Y')
            filtered_data = filtered_data[filtered_data[date_column] >= start_date]

        if end_date:
            end_date = datetime.strptime(end_date, '%d.%m.%Y')
            filtered_data = filtered_data[filtered_data[date_column] <= end_date]

        # Фильтрация по кварталам
        if quarter:
            filtered_data = filtered_data[filtered_data[date_column].dt.quarter == quarter]

        # Фильтрация по полугодиям
        if half_year:
            if half_year == 1:
                filtered_data = filtered_data[filtered_data[date_column].dt.month <= 6]
            elif half_year == 2:
                filtered_data = filtered_data[filtered_data[date_column].dt.month > 6]
        if drop_date:
            filtered_data = filtered_data.drop(drop_date, axis=1)

        new_dataframe = filtered_data.copy()

        # Переименование столбцов
        if dict_to_rename is not None:
            new_dataframe = new_dataframe.rename(columns=dict_to_rename)

        # Сохранение в Excel, если требуется
        if file_nam:
            with pd.ExcelWriter(file_nam, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                new_dataframe.to_excel(writer, sheet_name=sheet_nam, index=False)

        return new_dataframe


    except Exception as e:
        print(f"Произошла ошибка: {e}")


